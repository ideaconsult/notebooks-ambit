"""Dump sheet names, header rows, and a sample of data rows from an Excel
file, so a repair/authoring session can see the target layout without
opening the file itself.

Read-only -- never writes back to the Excel file. Dispatches on extension:
openpyxl for `.xlsx`/`.xlsm` (it cannot read the legacy binary format at
all -- confirmed live: raises InvalidFileException on a real `.xls`), xlrd
for `.xls` (many real nanodata-* spreadsheets, e.g. the NanoTest templates,
are still legacy `.xls`).
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import openpyxl
import xlrd


def _cell_str(value) -> str:
    if value is None:
        return ""
    return str(value)


@dataclass
class SheetPreview:
    name: str
    index: int          # 1-based, matching driver-properties '#N' syntax
    n_rows: int
    n_cols: int
    header_rows: list[list[str]]   # first up to `header_rows` rows, in full
    sample_rows: list[list[str]]   # next up to `sample_rows` rows, in full


@dataclass
class SpreadsheetPreview:
    path: Path
    sheet_names: list[str]
    sheets: list[SheetPreview]


def _resolve_targets(
    sheet: int | str | None, sheet_names: list[str]
) -> list[tuple[int, str]]:
    if sheet is None:
        return list(enumerate(sheet_names, start=1))
    if isinstance(sheet, int):
        return [(sheet, sheet_names[sheet - 1])]
    return [(sheet_names.index(sheet) + 1, sheet)]


def _read_xlsx(
    path: Path, sheet: int | str | None, header_rows: int, sample_rows: int,
    max_cols: int,
) -> SpreadsheetPreview:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_names = wb.sheetnames
        targets = _resolve_targets(sheet, sheet_names)

        previews = []
        for idx, name in targets:
            ws = wb[name]
            n_rows = ws.max_row or 0
            n_cols = ws.max_column or 0

            rows_wanted = header_rows + sample_rows
            row_iter = ws.iter_rows(
                max_row=rows_wanted, max_col=min(n_cols, max_cols)
            )
            collected: list[list[str]] = []
            for i, row in enumerate(row_iter, start=1):
                collected.append([_cell_str(c.value) for c in row])
                if i >= rows_wanted:
                    break

            sample_end = header_rows + sample_rows
            previews.append(
                SheetPreview(
                    name=name,
                    index=idx,
                    n_rows=n_rows,
                    n_cols=n_cols,
                    header_rows=collected[:header_rows],
                    sample_rows=collected[header_rows:sample_end],
                )
            )

        return SpreadsheetPreview(
            path=path, sheet_names=sheet_names, sheets=previews
        )
    finally:
        wb.close()


def _read_xls(
    path: Path, sheet: int | str | None, header_rows: int, sample_rows: int,
    max_cols: int,
) -> SpreadsheetPreview:
    wb = xlrd.open_workbook(str(path))
    sheet_names = wb.sheet_names()
    targets = _resolve_targets(sheet, sheet_names)

    previews = []
    for idx, name in targets:
        ws = wb.sheet_by_name(name)
        n_rows = ws.nrows
        n_cols = ws.ncols

        rows_wanted = min(header_rows + sample_rows, n_rows)
        cols_wanted = min(n_cols, max_cols)
        collected = [
            [_cell_str(ws.cell_value(r, c)) for c in range(cols_wanted)]
            for r in range(rows_wanted)
        ]

        sample_end = header_rows + sample_rows
        previews.append(
            SheetPreview(
                name=name,
                index=idx,
                n_rows=n_rows,
                n_cols=n_cols,
                header_rows=collected[:header_rows],
                sample_rows=collected[header_rows:sample_end],
            )
        )

    return SpreadsheetPreview(
        path=path, sheet_names=sheet_names, sheets=previews
    )


def read_spreadsheet(
    path: Path,
    sheet: int | str | None = None,
    *,
    header_rows: int = 3,
    sample_rows: int = 5,
    max_cols: int = 60,
) -> SpreadsheetPreview:
    """Preview one Excel file.

    `sheet` selects a single worksheet to preview in full: a 1-based index
    (matching the `#N` suffix convention in driver .properties lines) or a
    sheet name. If None, every sheet is previewed (cheap -- only the first
    `header_rows + sample_rows` rows of each are read).
    """
    reader = _read_xls if path.suffix.lower() == ".xls" else _read_xlsx
    return reader(path, sheet, header_rows, sample_rows, max_cols)
