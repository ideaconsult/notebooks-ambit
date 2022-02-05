# + tags=["parameters"]
upstream = ["onto"]
folder_output = None
model_embedding = None
hnsw_distance = None
terms_file = None
prefix = None
ann_index = None
spreadsheets = None
# -
import os,os.path
import pandas as pd
import torch, hnswlib
use_cuda = torch.cuda.is_available()
if use_cuda:
    print('__CUDNN VERSION:', torch.backends.cudnn.version())
    print('__Number CUDA Devices:', torch.cuda.device_count())
    print('__CUDA Device Name:', torch.cuda.get_device_name(0))
    print('__CUDA Device Total Memory [GB]:', torch.cuda.get_device_properties(0).total_memory / 1e9)
    torch_device = "cuda:0"
else:
    torch_device = "cpu"
from sentence_transformers import SentenceTransformer
    #distilbert-base-nli-stsb-mean-tokens for symmetric search

model = SentenceTransformer(model_embedding, device=torch_device)
e_idx = hnswlib.Index(space=hnsw_distance,dim=768)
e_idx.load_index(ann_index )
terms = pd.read_csv(terms_file,sep="\t",encoding="utf-8")

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
import os,os.path
import json


lookup = {}

def find_value(query,file, sheet, row,col, k=5):
    result = None
    if not query in lookup:
        result = []
        try:
            embeddings = model.encode(query, 
                                show_progress_bar=False,
                                normalize_embeddings=True)
            labels,distances =  e_idx.knn_query(embeddings, k=k)
            for label, distance in zip(labels[0],distances[0]):
                if distance<0.3:
                    tmp = {}
                    for tag in ["Class ID","Preferred Label","Definitions","ontology"]:
                        tmp[tag] = terms.iloc[label][tag]
                    result.append(tmp)
        except:
            pass
        lookup[query] = {"term": result} #, "where" : [(os.path.basename(file),sheet,row,col)] };
    else:
        result = lookup[query]    
        #lookup[query]["where"].append((os.path.basename(file),sheet,row,col))
    return result


                      
tmp = []
pathlist = Path(spreadsheets).glob('**/*.xlsx')
for path in pathlist:
    print(path)
    try:
        workbook = load_workbook(filename=str(path))
        for sheetname in workbook.sheetnames: 
            sheet = workbook[sheetname]
            print(sheetname)
            for row in sheet.rows:
                for cell in row:
                    if not (cell.value is None):
                        try:
                            float(cell.value)
                        except:
                            find_value(cell.value,path,sheetname, str(cell.row),cell.col_idx)
                            
        #break
        workbook.close()
        outfile = os.path.join(folder_output,"pykeen","lookup.json")
        with open(outfile, "w",encoding="utf-8") as write_file:
                json.dump(lookup, write_file)    
    except Exception as err:
        print(err)


